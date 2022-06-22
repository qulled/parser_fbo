from logging.handlers import RotatingFileHandler

from googleapiclient import discovery
from google.oauth2 import service_account
import requests
from googleapiclient.discovery import build
import logging
import os
import datetime as dt
import json

import openpyxl
import warnings


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(BASE_DIR, 'logs/')
log_file = os.path.join(BASE_DIR, 'logs/pars_table.log')
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler(
    log_file,
    maxBytes=100000,
    backupCount=3,
    encoding='utf-8'
)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s, %(levelname)s, %(message)s',
    handlers=(
        file_handler,
        console_handler
    )
)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

CREDENTIALS_FILE = 'credentials_service.json'
credentials = service_account.Credentials.from_service_account_file(CREDENTIALS_FILE)
service = discovery.build('sheets', 'v4', credentials=credentials)
START_POSITION_FOR_PLACE = 14


def get_excel(name_ip, token, date_from, date_to):
    url = 'https://suppliers-stats.wildberries.ru/api/v1/supplier/reportDetailByPeriod'
    params = {
        'key': token,
        'dateFrom': date_from,
        'dateto': date_to,
        'limit': 100000,
        'rrdid': 0
    }
    response = requests.get(url, params=params)
    with open(f'{date_to}-{name_ip}.json', 'w') as f:
        json.dump(response.json(), f, indent=2, ensure_ascii=False)
    return f'{date_to}-{name_ip}.json'


def dict_article_count(employees_sheet):
    dict_article_count = {}
    for x in range(3, employees_sheet.max_row + 1):
        if employees_sheet.cell(row=x, column=7).value not in dict_article_count and \
                employees_sheet.cell(row=x, column=11).value != 'Склад поставщика' and \
                employees_sheet.cell(row=x, column=13).value != 0:
            dict_article_count[employees_sheet.cell(row=x, column=7).value] = \
                employees_sheet.cell(row=x, column=13).value
        elif employees_sheet.cell(row=x, column=7).value in dict_article_count and \
                employees_sheet.cell(row=x, column=11).value != 'Склад поставщика' and \
                employees_sheet.cell(row=x, column=13).value != 0:
            dict_article_count[employees_sheet.cell(row=x, column=7).value] += \
                employees_sheet.cell(row=x, column=13).value
    return dict_article_count


def get_end_begining(day, month, year):
    return (f'{year}-{month}-{day}T23:59:59+03:00',
            f'{year}-{month}-{day}T00:00:00.00+03:00')


def convert_to_column_letter(column_number):
    column_letter = ''
    while column_number != 0:
        c = ((column_number - 1) % 26)
        column_letter = chr(c + 65) + column_letter
        column_number = (column_number - c) // 26
    return column_letter


def update_table_count_fbo(day, month, year, table_id, dict_count):
    global report_article
    range_name = f'{month}.{year}'
    position_for_place = START_POSITION_FOR_PLACE + (int(day) - 1) * 6

    service = build('sheets', 'v4', credentials=credentials)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=table_id,
                                range=range_name, majorDimension='ROWS').execute()

    values = result.get('values', [])
    i = 3
    body_data = []
    if not values:
        logging.info('No data found.')
    else:
        for row in values[2:]:
            try:
                article = int(row[7].strip().upper())
                if article in dict_count:
                    count = dict_count[article]
                    dict_count.pop(article)
                    if count == 0:
                        pass
                    body_data += [
                        {'range': f'{range_name}!{convert_to_column_letter(position_for_place + 2)}{i}',
                         'values': [[f'{count}']]}]
            except:
                pass
            finally:
                i += 1
                body = {
                    'valueInputOption': 'USER_ENTERED',
                    'data': body_data}
                global report_article
                if len(dict_count) > 0:
                    report_article = dict_count
    sheet.values().batchUpdate(spreadsheetId=table_id, body=body).execute()


if __name__ == '__main__':
    cred_file = os.path.join(BASE_DIR, 'credentials.json')

    day = dt.datetime.now().strftime("%d")
    month = dt.datetime.now().strftime("%m")
    year = dt.datetime.now().year

    with open(cred_file, 'r', encoding="utf-8") as f:
        cred = json.load(f)
        for i in cred:
            if i != 'Савельева':
                table_id = cred[i].get('table_id')
                with warnings.catch_warnings(record=True):
                    warnings.simplefilter("always")
                    excel_file = openpyxl.load_workbook(f'excel_docs/{i}-{dt.datetime.date(dt.datetime.now())}.xlsx')
                employees_sheet = excel_file['Sheet1']
                update_table_count_fbo(day, month, year, table_id, dict_article_count(employees_sheet))
                if i == 'Кулик':
                   table_id = cred[i].get('table_id_december')
                with warnings.catch_warnings(record=True):
                    warnings.simplefilter("always")
                    excel_file = openpyxl.load_workbook(f'excel_docs/{i}-{dt.datetime.date(dt.datetime.now())}.xlsx')
                employees_sheet = excel_file['Sheet1']
                update_table_count_fbo(day, month, year, table_id, dict_article_count(employees_sheet))
